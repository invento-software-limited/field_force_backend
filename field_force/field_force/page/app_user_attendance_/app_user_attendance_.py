import copy
import io
import os

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import frappe
import json
import openpyxl

from field_force.field_force.report.utils import set_image_url, set_user_link, get_site_directory_path

export_data = []

@frappe.whitelist()
def get_user_attendance_data(filters=None):
    global export_data
    columns = get_columns()
    query_result = get_query_data(filters)
    site_directory = get_site_directory_path()
    export_data = copy.deepcopy(query_result)

    for index, app_user_attendance in enumerate(query_result):
        # generate_row_export_data(index, app_user_attendance)
        set_user_link(app_user_attendance)

        app_user_attendance['name'] = f'<a href="/app/app-user-attendance/{app_user_attendance.name}" ' \
                                      f'target="_blank">{app_user_attendance.name}</a>'
        app_user_attendance.server_date = f"{app_user_attendance.server_date}<br>{app_user_attendance.server_time}"
        app_user_attendance.device_date = f"{app_user_attendance.device_date}<br>{app_user_attendance.device_time}"

        if app_user_attendance.cheated:
            app_user_attendance.cheated = 'Yes'

        set_image_url(app_user_attendance, site_directory)
        app_user_attendance['sl'] = index + 1

    # data = query_result
    return query_result, columns

def generate_row_export_data(index, row):
    export_data[index].server_date = f"{row.server_date}\n{row.server_time}"
    export_data[index].device_date = f"{row.device_date}\n{row.device_time}"
    export_data[index].cheated = 'Yes' if export_data[index].cheated else ''
    export_data[index]['sl'] = index + 1

def get_query_data(filters):
    try:
        filters = json.loads(filters)
    except:
        pass
    conditions = get_conditions(filters)

    query_string = """select app_user_attendance.name, app_user_attendance.user, app_user_attendance.user_fullname, 
                    app_user_attendance.server_date, app_user_attendance.server_time, app_user_attendance.type,
                    app_user_attendance.device_date, app_user_attendance.device_time, app_user_attendance.latitude,
                    app_user_attendance.longitude, app_user_attendance.device_model, app_user_attendance.image,
                    app_user_attendance.cheated from `tabApp User Attendance` app_user_attendance where %s
                    order by app_user_attendance.server_date desc""" % (conditions)

    query_result = frappe.db.sql(query_string, as_dict=1, debug=0)
    return query_result


def get_conditions(filters):
    from_date = filters.get('from_date')
    to_date = filters.get('to_date')
    user = filters.get('user')
    type = filters.get('type')

    conditions = []

    if type:
        conditions.append("app_user_attendance.type = '%s'" % type)
    if from_date:
        conditions.append("app_user_attendance.server_date >= '%s'" % from_date)
    if to_date:
        conditions.append("app_user_attendance.server_date <= '%s'" % to_date)
    if user:
        conditions.append("app_user_attendance.user = '%s'" % user)

    return " and ".join(conditions)

def get_columns():
    columns =  [
        {'fieldname': 'sl', 'label': 'SL', 'expwidth': 5},
        {'fieldname': 'server_date', 'label': 'DateTime', 'expwidth': 15},
        {'fieldname': 'name', 'label': 'ID', 'expwidth': 15},
        {'fieldname': 'user', 'label': 'User', 'expwidth': 15},
        {'fieldname': 'type', 'label': 'Type', 'expwidth': 15},
        {'fieldname': 'device_date', 'label': 'Device DateTime', 'expwidth': 15},
        {'fieldname': 'cheated', 'label': 'Cheated', 'expwidth': 15},
        {'fieldname': 'latitude', 'fieldtype': 'Currency', 'label': 'Latitude', 'expwidth': 15},
        {'fieldname': 'longitude', 'label': 'Longitude', 'expwidth': 15},
        {'fieldname': 'device_model', 'label': 'Model', 'expwidth': 15},
        {'fieldname': 'image', 'label': 'Image', 'fieldtype': 'Image', 'expwidth': 15, 'export': False}
    ]
    return columns

@frappe.whitelist()
def export_data():
    columns = get_columns()
    generate_excel_and_download(columns, export_data)

def generate_excel_and_download(columns, data):
    fields, labels, row_widths= [], [], []

    for field in columns:
        if field.get('export', True):
            fields.append(field.get('fieldname'))
            labels.append(field.get('label'))
            row_widths.append(field.get('expwidth', 25))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    generate_row(worksheet, 1, labels, height=25)
    row_count = 2

    for row in data:
        row_data = [row[field] for field in fields]
        generate_row(worksheet, row_count, row_data, widths=row_widths, height=35)
        row_count += 1

    byte_data = io.BytesIO()
    workbook.save(byte_data)

    frappe.local.response.filename = 'App_User_Attendance_Report.xlsx'
    frappe.local.response.filecontent = byte_data.getvalue()
    frappe.local.response.type = "download"


def generate_row(ws, row_count, column_values, font=None, font_size=None, color=None, height=25, widths=None):
    cells = []

    for i, value in enumerate(column_values):
        column_number = i + 1
        cell = ws.cell(row=row_count, column=column_number)
        cell.value = value

        if font:
            cell.font = font
        elif font_size:
            cell.font = Font(size=font_size)

        if color:
            cell.fill = PatternFill(fgColor=color, fill_type='solid')
        if widths:
            ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

        if isinstance(value, int):
            cell.number_format = "#,##0"
        elif isinstance(value, float):
            cell.number_format = "#,##0.00"

        cell.alignment = Alignment(vertical='center')
        cells.append(cell)

    if height:
        ws.row_dimensions[row_count].height = height

    return cells