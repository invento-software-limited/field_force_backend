import datetime
import io
import frappe
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel_and_download(columns, data, file_name, height=25):
    fields, labels, row_widths= [], [], []

    for field in columns:
        if field.get('export', True):
            fields.append(field.get('fieldname'))
            labels.append(field.get('label'))
            row_widths.append(field.get('expwidth', 15))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    generate_row(worksheet, 1, labels, height=20)
    row_count = 2

    for row in data:
        row_data = [row.get(field) for field in fields]
        generate_row(worksheet, row_count, row_data, widths=row_widths, height=height)
        row_count += 1

    byte_data = io.BytesIO()
    workbook.save(byte_data)

    frappe.local.response.filename = file_name
    frappe.local.response.filecontent = byte_data.getvalue()
    frappe.local.response.type = "download"
    return frappe.local.response


def generate_row(ws, row_count, column_values, font=None, font_size=None, color=None, height=25, widths=None):
    cells = []

    for i, value in enumerate(column_values):
        column_number = i + 1
        cell = ws.cell(row=row_count, column=column_number)
        cell.value = value

        if font:
            cell.font = font
        elif font_size:
            cell.font = Font(size=font_size)

        if color:
            cell.fill = PatternFill(fgColor=color, fill_type='solid')
        if widths:
            ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]

        if isinstance(value, int):
            cell.number_format = "#,##0"
        elif isinstance(value, float):
            cell.number_format = "#,##0.00"

        cell.alignment = Alignment(vertical='center')
        cells.append(cell)

    if height:
        ws.row_dimensions[row_count].height = height

    return cells

def get_time_in_12_hour_format(time):
    if time and time != 'None':
        time = datetime.datetime.strptime(str(time), '%H:%M:%S').time()

    if isinstance(time, datetime.time):
        return time.strftime('%l:%M %p')

    return time

def get_datetime_with_12_hour_format(datetime_):
    if datetime_:
        datetime_ = datetime.datetime.strptime(datetime_, '%Y-%m-%d %H:%M:%S')
        return datetime_.strftime('%Y-%m-%d %l:%M %p')

    return datetime_

def get_doc_url(doc_url, name, label=None):
    return f'<a href="/app/{doc_url}/{name}"' f'target="_blank">' \
           f'<span style="color: blue;">{label}<span></a>'

def get_spent_time(in_time, out_time, in_word=True):
    if not in_time and not out_time:
        return ''

    if isinstance(in_time, str):
        in_time = get_timedelta_time_obj(in_time)
    if isinstance(out_time, str):
        out_time = get_timedelta_time_obj(out_time)

    spent_time = out_time - in_time

    if in_word:
        time_elements = str(spent_time).split(':')
        hour, minutes = int(time_elements[0]), int(time_elements[1])
        spent_time = ''

        if hour:
            spent_time = f"{hour} hours" if hour > 1 else f"{hour} hour"
        if minutes:
            if spent_time:
                spent_time += " and"

            spent_time += f" {minutes} minutes" if minutes > 1 else f"{minutes} minute"

    return spent_time

def get_timedelta_time_obj(time_str):
    out_time = datetime.datetime.strptime(str(time_str), '%H:%M:%S')
    return datetime.timedelta(hours=out_time.hour, minutes=out_time.minute,
                              seconds=out_time.second, microseconds=out_time.microsecond)
