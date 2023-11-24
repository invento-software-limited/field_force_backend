# Copyright (c) 2022, Invento Software Limited and contributors
# For license information, please see license.txt
import calendar
import copy
import os
import random
from frappe.model.mapper import get_mapped_doc

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import frappe
import time
import datetime
from frappe.model.document import Document
from field_force.field_force.doctype.utils import get_data_from_pdf

import openpyxl

class Requisition(Document):
    def validate(self):
        self.set_user()
        self.set_customer_info()
        self.set_partner_group()
        self.set_brand_and_image_to_requisition_items()
        self.validate_delivery_date()
        # self.validate_po_number()
        self.validate_items()

    def before_save(self):
        set_extra_values(self)
        add_sales_person(self)
        add_amount_to_achievement(self)
        set_allocated_amount(self)
        # generate_requisition_excel_and_attach(self)

    # def before_submit(self):
    #     generate_csv_and_attach_file(self)
        # generate_requisition_excel_and_attach(self)

    def on_submit(self):
        if self.docstatus == 1:
            frappe.db.set_value(self.doctype, self.name, 'status', 'Submitted')

        month, year = get_month_and_year(self)
        set_amount_to_sales_target(self.sales_person, self.achievement_amount, month, year)
        create_sales_order_on_submit(self)

    def on_cancel(self):
        subtract_achievement_amount(self)

        if self.docstatus == 2:
            frappe.db.set_value(self.doctype, self.name, 'status', 'Cancelled')

    def validate_delivery_date(self):
        if self.transaction_date > self.delivery_date:
            frappe.throw("Expected delivery date should be after Requisition date")

        for item in self.items:
            if not item.delivery_date:
                item.delivery_date = self.delivery_date

    def validate_po_number(self):
        filters =  {
            "po_no": self.po_no,
            "name": ["!=", self.name],
            "docstatus": ["!=", 2]
        }

        if frappe.db.exists("Requisition", filters):
            frappe.throw(f"Requisition already exists with the PO number(<b>{self.po_no}</b>)", title="Already Exist")

    def set_user(self):
        if not self.user:
            self.user = frappe.session.user
            self.user_fullname = frappe.db.get_value('User', self.user, 'full_name')

    def set_customer_info(self):
        if not self.customer_name or not self.distributor:
            customer_name, distributor, customer_group = frappe.db.get_value('Customer', self.customer,
                                                                             ['customer_name', 'distributor', 'customer_group'])
            if not self.customer_name and customer_name:
                self.customer_name = customer_name

            if not self.distributor and distributor:
                self.distributor = distributor

            if not self.customer_group and customer_group:
                self.customer_group = customer_group

    def set_partner_group(self):
        if not self.partner_group:
            self.partner_group = frappe.db.get_value('Customer', {'name': self.customer}, 'partner_group')

    def set_brand_and_image_to_requisition_items(self):
        for item in self.items:
            if not item.brand or not item.image:
                try:
                    item.brand, item.image = frappe.db.get_value('Item', item.item_code, ['brand', 'image'])
                except:
                    pass

    def validate_items(self):
        commission_brand_dict = get_brands_commission(self.customer)
        total_items = 0
        total_qty = 0
        total_accepted_qty = 0
        total_accepted_amount = 0
        total = 0

        if self.items:
            for item in self.items:
                if not item.qty:
                    frappe.throw(f"Please give quantity of item <b>{item.item_name}</b>")

                if not item.discount_percentage:
                    item.discount_percentage = commission_brand_dict.get(item.brand, 0)

                if not item.price_list_rate:
                    item.price_list_rate = frappe.db.get_value("Item Price", {"item_code": item.item_code,
                                                                              "selling": 1}, ["price_list_rate"])

                if not item.price_list_rate and not item.rate:
                    frappe.throw(f"There is no selling rate of item <b>{item.item_name}</b>")

                elif item.price_list_rate or item.rate:
                    if (item.price_list_rate and not item.rate) or (item.price_list_rate == item.rate):
                        if item.discount_percentage:
                            item.discount_amount = item.price_list_rate * (float(item.discount_percentage) / 100)
                        else:
                            item.discount_amount = 0

                        item.rate = item.price_list_rate - item.discount_amount

                    item.amount = item.qty * item.rate
                    total += float(item.amount)

                if item.accepted_qty:
                    item.accepted_amount = item.accepted_qty * item.rate
                    total_accepted_qty += float(item.accepted_qty)
                    total_accepted_amount += float(item.accepted_amount)

                total_items += 1
                total_qty += int(item.qty)

                # self.validate_accepted_qty(item)
        else:
            frappe.throw("Items are required")

        if total:
            self.total = total
            self.net_total = total
            self.grand_total = total

        if total_items and total_qty:
            self.total_qty = total_qty
            self.total_items = total_items

        if total_accepted_qty:
            self.total_accepted_qty = total_accepted_qty
            self.total_accepted_amount = total_accepted_amount

    def validate_accepted_qty(self, item):
        if not item.accepted_qty:
            item.accepted_qty = item.qty

def set_extra_values(self):
    self.total_items = len(self.items)

    for item in self.items:
        if not item.brand or not item.image:
            try:
                item.brand, item.image = frappe.db.get_value('Item', item.item_code, ['brand', 'image'])
            except:
                pass

def set_allocated_amount(self):
    if self.sales_team:
        for sales_person in self.sales_team:
            if self.total and sales_person.allocated_percentage:
                sales_person.allocated_amount = (self.total * sales_person.allocated_percentage)/100

def add_sales_person(self):
    if not self.sales_person and frappe.db.exists("Sales Person", {"user": self.owner}):
        self.sales_person, self.employee, self.user = frappe.db.get_value("Sales Person", {'user': self.owner},
                                                               ['name', 'employee', 'user'])
    if not self.sales_team and self.sales_person:
        self.append("sales_team", {
            "parent": self.name,
            "parenttype": "Sales Order",
            "sales_person": self.sales_person,
            "allocated_percentage": 100,
            "allocated_amount": self.grand_total,
            "employee": self.employee,
            "user": self.user
        })

    if not self.user:
        self.user = self.owner

def add_amount_to_achievement(self):
    achievement_amount = get_default(self.grand_total) - get_default(self.achievement_amount)

    if self.achievement_amount or self.achievement_amount == 0:
        self.achievement_amount += achievement_amount
    else:
        self.achievement_amount = achievement_amount

def subtract_achievement_amount(self):
    month, year = get_month_and_year(self)
    achievement_amount = -self.achievement_amount
    set_amount_to_sales_target(self.sales_person, achievement_amount, month, year)

def set_amount_to_sales_target(sales_person, achievement_amount, month, year):
    filters = {
        "sales_person": sales_person,
        "month": calendar.month_name[month],
        "year": year,
        "docstatus": 1
    }

    if frappe.db.exists("Sales Person Target", filters):
        sales_target = frappe.get_doc("Sales Person Target", filters)
        sales_target.achievement_amount += achievement_amount
        sales_target.save()

def get_default(value):
    if value:
        return value
    return 0

def get_month_and_year(sales_order):
    year = frappe.defaults.get_user_default('fiscal_year')

    if isinstance(sales_order.transaction_date, datetime.date):
        return sales_order.transaction_date.month, year
    elif isinstance(sales_order.transaction_date, str):
        month =  datetime.datetime.strptime(sales_order.transaction_date, "%Y-%m-%d").month
        return month, year
    return None, None

def generate_requisition_excel_and_attach(requisition):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.name = requisition.name

    columns = ['[Product #]', '[Product Name]', '[UOM]', '[Quantity]', '[Unit Price]', '[Discount]', '[Nett Price]',
               '[Amount]', '[Tax Code]']

    generate_row(worksheet, 1, columns, font=default_font)
    row_count = 2

    for item in requisition.items:
        item_row_data = [item.item_code, item.item_name, item.uom, item.qty, item.price_list_rate,
                         item.discount_amount, f'=E{row_count}-F{row_count}',
                         f'=D{row_count}*G{row_count}', 'NA']

        generate_row(worksheet, row_count, item_row_data, font=default_font)
        row_count += 1

    file_name = f"Requisition_{requisition.name}.xlsx"
    file_path = get_directory_path('requisition/')
    absolute_path = file_path + file_name
    worksheet.sheet_view.zoomScale = 90
    workbook.save(absolute_path)

    file = attach_file(requisition, absolute_path, file_name)
    requisition.requisition_excel = file.file_url
    requisition.requisition_excel_file = f'<a class="attached-file-link" href="{file.file_url}"' \
                                         f' target="_blank">{file_name}</a>'
        # + f"?file={random.randint(100, 1000)}"

def attach_file(requisition, file_path, file_name):
    with open(file_path, "rb") as data:
        content = data.read()

    file = frappe.get_doc({
        "doctype": "File",
        "attached_to_doctype": requisition.get("doctype"),
        "attached_to_name": requisition.get("name"),
        "attached_to_field": 'requisition_excel',
        "folder": 'Home/Attachments',
        "file_name": file_name,
        "file_url": '',
        "is_private": 0,
        "content": content
    })
    file.save()
    return file

def generate_row(ws, row_count, column_values, font=None, font_size=None, color=None, height=None):
    cells = []
    amount_columns = [5, 7, 8]
    column_widths = [15, 15, 15,15, 15, 15,15, 15, 15,]

    for i, value in enumerate(column_values):
        column_number = i + 1
        cell = ws.cell(row=row_count, column=column_number)
        cell.value = value

        if font:
            cell.font = font
        elif font_size:
            cell.font = Font(size=font_size)
        if color:
            cell.fill = PatternFill(fgColor=color, fill_type='solid')

        # if isinstance(value, int):
        #     cell.number_format = "#,##0"
        # elif isinstance(value, float):
        if i+1 in amount_columns:
            cell.number_format = "#,##0.00"

        ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i]
        cell.alignment = Alignment(vertical='center')
        cells.append(cell)

    if height:
        ws.row_dimensions[row_count].height = height

    set_column_width(ws)
    return cells

def get_directory_path(directory_name):
    site_name = frappe.local.site
    cur_dir = os.getcwd()
    file_path = os.path.join(cur_dir, site_name, f'public/files/{directory_name}')
    os.makedirs(file_path, exist_ok=True)
    return file_path

default_font = Font(name='Calibri(Body)', size=11, bold=False, italic   =False,  strike=False, underline='none')

def set_column_width(worksheet, column=None, width=None):
    if column and width:
        worksheet.column_dimensions[get_column_letter(column)].width = width
    else:
        for i, width_ in enumerate([12, 15, 12, 12, 12, 15]):
            worksheet.column_dimensions[get_column_letter(i + 1)].width = width_

@frappe.whitelist()
def get_brands_commission(customer, brand=None):
    customer_group, distributor = frappe.db.get_value("Customer", customer, ['customer_group', 'distributor'])
    brand_wise_commission_dict = {}

    if customer_group == "Retail Shop" and distributor:
        commissions = frappe.db.get_all('Distributor Brand Commission', {'parent': distributor},
                                        ['brand', 'commission_rate'])
        brand_wise_commission_dict = get_commissions_in_dict(commissions)

    if not brand_wise_commission_dict:
        commissions = frappe.db.get_all('Customer Brand Commission', {'parent': customer},
                                        ['brand', 'commission_rate'])
        brand_wise_commission_dict = get_commissions_in_dict(commissions)

    if brand:
        return brand_wise_commission_dict.get(brand, 0)

    return brand_wise_commission_dict

@frappe.whitelist()
def get_item_details(item_code, fields=None):
    fields = fields or ['name as item_code', 'product_id', 'item_name', 'brand']
    item = {}

    if item_code:
        if frappe.db.exists("Item", {'item_code': item_code, 'disabled': 0}):
            item = frappe.db.get_value("Item", item_code, fields , as_dict=1)
        elif frappe.db.exists("Item", {'product_id': item_code, 'disabled': 0}):
            item = frappe.db.get_value("Item", {'product_id': item_code, 'disabled': 0}, fields, as_dict=1)
        else:
            item = get_item_by_barcode(item_code)

    if item:
        try:
            item.price_list_rate = frappe.db.get_value("Item Price", {'item_code': item.item_code,
                                                                      'selling':1}, 'price_list_rate')
        except:
            item.price_list_rate = 0

    return item

def get_item_by_barcode(barcode):
    items = frappe.db.sql('''select barcode.barcode, item.item_code, item.item_name, item.product_id, item.brand
                            from `tabItem Barcode` barcode join `tabItem` item on barcode.parent=item.item_code
                            where barcode.barcode="%s" and item.disabled=0''' % barcode, as_dict=1)
    if items:
        return items[0]
    return {}


def get_commissions_in_dict(commissions):
    brand_wise_commission_dict = {}

    for commission in commissions:
        brand_wise_commission_dict[commission.brand] = commission.commission_rate

    return brand_wise_commission_dict

import csv
import json

@frappe.whitelist()
def generate_csv_and_attach_file(requisition):
    try:
        requisition = json.loads(requisition)
    except:
        pass

    file_name = f"Requisition_{requisition.get('name')}.csv"
    file_path = get_directory_path('requisition/')
    absolute_path = file_path + file_name

    columns = ['Product #', 'Product Name', 'UOM', 'Quantity', 'Unit Price', 'Discount', 'Nett Price', 'Amount']

    with open(absolute_path, 'w+', encoding='UTF8') as f:
        writer = csv.writer(f)

        # write the header
        writer.writerow(columns)

        for item in requisition.get("items"):
            item_row_data = [item.get("product_id"), item.get("item_name"),
                             item.get("uom"), item.get("qty"), item.get("rate"), '', item.get("rate"), item.get("amount")]
            # write the data
            writer.writerow(item_row_data)


    attach_file(requisition, absolute_path, file_name)
    file_url = absolute_path.split('public')[-1]

    return {
        "url" : file_url,
        "name" : file_name
    }

@frappe.whitelist()
def update_requisition_from_csv_file(file_url):
    file_path = get_site_directory_path() + "/public/" + file_url
    data = []
    with open(file_path, 'r') as csvfile:
        csvreader = csv.DictReader(csvfile)
        for row in csvreader:
            data.append(row)

    return data

def create_sales_order_on_submit(requisition):
    if requisition.create_sales_order_on_submit:
        sales_order = frappe.get_doc(frappe._dict({
            "doctype": "Sales Order",
            "customer": requisition.customer,
            "transaction_date": requisition.transaction_date,
            "delivery_date": requisition.delivery_date,
            "order_type": "Sales",
        }))

        for item in requisition.items:
            sales_order.append("items", {
                "item_code": item.item_code,
                "item_name": item.item_name,
                "brand": item.brand,
                "qty": item.accepted_qty if item.accepted_qty else item.qty,
                "rate": item.rate,
                "amount": (item.rate * item.accepted_qty) if item.accepted_qty else (item.rate * item.qty)
            })

        sales_order.insert()

def get_site_directory_path():
    site_name = frappe.local.site
    cur_dir = os.getcwd()
    return os.path.join(cur_dir, site_name)


@frappe.whitelist()
def make_delivery_trip(source_name, target_doc=None):
    def update_stop_details(source_doc, target_doc, source_parent):
        target_doc.customer = source_parent.customer
        target_doc.address = source_parent.address
        target_doc.grand_total = source_parent.grand_total

        # Append unique Delivery Notes in Delivery Trip
        requisition.append(target_doc.delivery_note)

    requisition = []

    doclist = get_mapped_doc(
        "Requisition",
        source_name,
        {
            "Requisition": {"doctype": "Delivery Trip", "validation": {"docstatus": ["=", 1]}},
            "Requisition Item": {
                "doctype": "Delivery Stop",
                "field_map": {"parent": "requisition"},
                "condition": lambda item: item.parent not in requisition,
                "postprocess": update_stop_details,
            },
        },
        target_doc,
    )

    return doclist
